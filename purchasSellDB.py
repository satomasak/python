#伝票作成システムからエクスポートしたデータをもとに仕入商品ごとの売り上げデータベースを作成する

import pandas as pd
#データファイル読み込み用のフォルダパス
inputFolderPath = 'C:\\Users\\sato\\Desktop\\purchaseData\\'
#データ出力用のフォルダパス
outputFolderPath = 'C:\\Users\\sato\\Desktop\\sellData\\'
################売上DB作成#####################################
#selldfに必要なカラム番号だけのデータフレームを売上.xlsから抽出する
selldf = pd.read_excel(inputFolderPath+'売上.xls',usecols=[1,5,6,9,11,12,15,17,18])
#欠損値を’なし’で置き換える
selldf = selldf.fillna('なし')

#procuctRistの作成　商品一覧　重複を削除し、ユニークデーターを抽出
sellProductList = selldf[['取引区分','商品名']].drop_duplicates()
#purchaseRistの作成　仕入れ一覧　取引区分の「掛その他」で抽出
sellPurchaseList = sellProductList[sellProductList['取引区分'] == '掛その他']
#仕入れ商品一覧の中からトマト、レタスを区分
tom = sellPurchaseList[sellPurchaseList['商品名'].str.contains('トマト|日向|ｷｬﾛﾙ|桃姫|ﾄﾏﾄ')]
let = sellPurchaseList[sellPurchaseList['商品名'].str.contains('レタス|ﾚﾀｽ')]


#返品以外のトマト商品のリスト
x = tom [~tom ['商品名'].str.contains('返品')]
sell_uniqueTomato = x['商品名']
#返品以外のレタス仕入商品のリスト
x = let [~let ['商品名'].str.contains('返品')]
sell_uniqueLettuce = x['商品名']

#返品リスト
x = tom[tom ['商品名'].str.contains('返品')]
sell_returne_uniqueTomato = x['商品名']
#返品以外のレタス仕入商品のリスト
x = let [let ['商品名'].str.contains('返品')]
sell_returne_uniqueLettuce = x['商品名']



#トマト仕入れ商品の売上データを商品名、得意先、売上単価ごとにグループ化する(返品以外)
selldf_tomato_array = []
for i in sell_uniqueTomato:
    nam = selldf[selldf['商品名']== i]
    #得意先コード、得意先名、商品名、売上単価をキーにして、同じ売上・数量は合計する
    nam = nam.drop(columns=['商品名', '取引区分'])
    nam = nam.groupby(['商品ｺｰﾄﾞ','取引年月日','売上単価','得意先ｺｰﾄﾞ','得意先名']).sum()
    selldf_tomato_array.append(nam)

#レタスの売上データを商品名、得意先、売上単価ごとにグループ分けする(返品以外)
selldf_lettuce_array = []
for i in sell_uniqueLettuce:
    nam = selldf[selldf['商品名']== i]
    nam = nam.drop(columns=['商品名', '取引区分'])
    nam = nam.groupby(['商品ｺｰﾄﾞ','取引年月日','売上単価','得意先ｺｰﾄﾞ','得意先名']).sum()
    selldf_lettuce_array.append(nam)

#トマト売上返品リストの作成
selldf_returneTomato_array =[]
for i in sell_returne_uniqueTomato:
    nam = selldf[selldf['商品名']== i]
    #得意先コード、得意先名、商品名、売上単価をキーにして、同じ売上・数量は合計する
    nam = nam.drop(columns=['商品名', '取引区分'])
    nam = nam.groupby(['商品ｺｰﾄﾞ','取引年月日','売上単価','得意先ｺｰﾄﾞ','得意先名']).sum()
    selldf_returneTomato_array.append(nam)

#仕入れレタス売上返品リストの作成
selldf_returneLettuce_array =[]
for i in sell_returne_uniqueLettuce:
    nam = selldf[selldf['商品名']== i]
    #得意先コード、得意先名、商品名、売上単価をキーにして、同じ売上・数量は合計する
    nam = nam.drop(columns=['商品名', '取引区分'])
    nam = nam.groupby(['商品ｺｰﾄﾞ','取引年月日','売上単価','得意先ｺｰﾄﾞ','得意先名']).sum()
    selldf_returneLettuce_array.append(nam)


#売上データDB作成
import openpyxl
import datetime
day = datetime.date.today()
dayName = day.strftime('%Y%m%d')

#トマト販売データベース
fileName = dayName + 'トマト販売データ.xlsx'
filePath = outputFolderPath+fileName
wb = openpyxl.Workbook()
for i in sell_uniqueTomato:
    wb.create_sheet(i)
wb.save(filePath)
j = 0
with pd.ExcelWriter(filePath, engine='openpyxl') as writer:
    for i in sell_uniqueTomato:
        selldf_tomato_array[j].to_excel(writer, sheet_name=i)
        j += 1

#売上列の3桁カンマ区切り
wb = openpyxl.load_workbook(filePath)
for i in sell_uniqueTomato:
    sh = wb[i]
    a = sh.max_row
    for b in range(a):
        sh.cell(row=b+1,column=7).number_format = '#,##0'
wb.save(filePath)



#レタス販売データベース
fileName = dayName + '仕入レタス販売データ.xlsx'
filePath = outputFolderPath+fileName
wb = openpyxl.Workbook()
for i in sell_uniqueLettuce:
    wb.create_sheet(i)
wb.save(filePath)
j = 0
with pd.ExcelWriter(filePath, engine='openpyxl') as writer:
    for i in sell_uniqueLettuce:
        selldf_lettuce_array[j].to_excel(writer, sheet_name=i)
        j += 1
#売上列の3桁カンマ区切り
wb = openpyxl.load_workbook(filePath)
for i in sell_uniqueLettuce:
    sh = wb[i]
    a = sh.max_row
    for b in range(a):
        sh.cell(row=b+1,column=7).number_format = '#,##0'
wb.save(filePath)


#トマト売上返品リストの作成
fileName = dayName + 'トマト売上返品データ.xlsx'
filePath = outputFolderPath+fileName
wb = openpyxl.Workbook()
for i in sell_returne_uniqueTomato:
    wb.create_sheet(i)
wb.save(filePath)
j = 0
with pd.ExcelWriter(filePath, engine='openpyxl') as writer:
    for i in sell_returne_uniqueTomato:
        selldf_returneTomato_array[j].to_excel(writer, sheet_name=i)
        j += 1

#売上列の3桁カンマ区切り
wb = openpyxl.load_workbook(filePath)
for i in sell_returne_uniqueTomato:
    sh = wb[i]
    a = sh.max_row
    for b in range(a):
        sh.cell(row=b+1,column=7).number_format = '#,##0'
wb.save(filePath)

#レタス売上返品データベース
fileName = dayName + '仕入レタス売上返品データ.xlsx'
filePath = outputFolderPath+fileName
wb = openpyxl.Workbook()
for i in sell_returne_uniqueLettuce:
    wb.create_sheet(i)
wb.save(filePath)
j = 0
with pd.ExcelWriter(filePath, engine='openpyxl') as writer:
    for i in sell_returne_uniqueLettuce:
        selldf_returneLettuce_array[j].to_excel(writer, sheet_name=i)
        j += 1

#売上列の3桁カンマ区切り
wb = openpyxl.load_workbook(filePath)
for i in sell_returne_uniqueLettuce:
    sh = wb[i]
    a = sh.max_row
    for b in range(a):
        sh.cell(row=b+1,column=9).number_format = '#,##0'
wb.save(filePath)
print('販売データ作成')

#######################売上DB作成以上####################



#######################仕入DB作成##########################
buydf = pd.read_excel(inputFolderPath+'仕入れ.xls',usecols=[1,5,6,9,11,12,15,17,22], skip_blank_lines=False)
#欠損値を’なし’で置き換える
buydf = buydf.fillna('なし')

#仕入procuctRistの作成　商品一覧　重複を削除し、ユニークデーターを抽出
buyProductList = buydf[['取引区分','商品名']].drop_duplicates()
#仕入れ一覧の中からトマト、レタスを区分
tom = buyProductList[buyProductList['商品名'].str.contains('トマト|日向|ｷｬﾛﾙ|桃姫|ﾄﾏﾄ')]
let = buyProductList[buyProductList['商品名'].str.contains('レタス|ﾚﾀｽ')]


#返品以外のトマト商品のリスト
x = tom [~tom ['商品名'].str.contains('返品')]
buy_uniqueTomato = x['商品名']
#返品以外のレタス仕入商品のリスト
x = let [~let ['商品名'].str.contains('返品')]
buy_uniqueLettuce = x['商品名']

#返品リスト
x = tom [tom ['商品名'].str.contains('返品')]
buy_returneTomato = x['商品名']
#返品以外のレタス仕入商品のリスト
x = let [let ['商品名'].str.contains('返品')]
buy_returneLettuce = x['商品名']


#トマト仕入れ商品の仕入データを商品名、得意先、売上単価ごとにグループ化する(返品以外)
buydf_tomao_array = []
for i in buy_uniqueTomato:
    nam = buydf[buydf['商品名']== i]
    #得意先コード、得意先名、商品名、売上単価をキーにして、同じ売上・数量は合計する 商品名と取引区分は削除
    nam = nam.drop(columns=['商品名', '取引区分'])
    nam = nam.groupby(['商品ｺｰﾄﾞ','取引年月日','仕入単価','仕入先ｺｰﾄﾞ','仕入先名','行摘要']).sum()
    buydf_tomao_array.append(nam)

#仕入レタスの仕入れデータを商品名、得意先、売上単価ごとにグループ分けする(返品以外)
buydf_lettuce_array = []
for i in buy_uniqueLettuce:
    nam = buydf[buydf['商品名']== i]
    nam = nam.drop(columns=['商品名', '取引区分'])
    nam = nam.groupby(['商品ｺｰﾄﾞ','取引年月日','仕入単価','仕入先ｺｰﾄﾞ','仕入先名','行摘要']).sum()
    buydf_lettuce_array.append(nam)

#トマト売上返品リストの作成
buydf_returneTomato_array =[]
for i in buy_returneTomato:
    nam = buydf[buydf['商品名']== i]
    #得意先コード、得意先名、商品名、売上単価をキーにして、同じ売上・数量は合計する
    nam = nam.drop(columns=['商品名', '取引区分'])
    nam = nam.groupby(['商品ｺｰﾄﾞ','取引年月日','仕入単価','仕入先ｺｰﾄﾞ','仕入先名','行摘要']).sum()
    buydf_returneTomato_array.append(nam)

#仕入れレタス売上返品リストの作成
buydf_returneLettuce_array =[]
for i in buy_returneLettuce:
    nam = buydf[buydf['商品名']== i]
    #得意先コード、得意先名、商品名、売上単価をキーにして、同じ売上・数量は合計する
    nam = nam.drop(columns=['商品名', '取引区分'])
    nam = nam.groupby(['商品ｺｰﾄﾞ','取引年月日','仕入単価','仕入先ｺｰﾄﾞ','仕入先名','行摘要']).sum()
    buydf_returneLettuce_array.append(nam)


#仕入DB作成
#トマト仕入データベース
fileName = dayName + 'トマト仕入データ.xlsx'
filePath = outputFolderPath+fileName
wb = openpyxl.Workbook()
for i in buy_uniqueTomato:
    wb.create_sheet(i)
wb.save(filePath)
j = 0
with pd.ExcelWriter(filePath, engine='openpyxl') as writer:
    for i in buy_uniqueTomato:
        buydf_tomao_array[j].to_excel(writer, sheet_name=i)
        j += 1



#レタス仕入データベース
fileName = dayName + '仕入レタス仕入データ.xlsx'
filePath = outputFolderPath+fileName
wb = openpyxl.Workbook()
for i in buy_uniqueLettuce:
    wb.create_sheet(i)
wb.save(filePath)
j = 0
with pd.ExcelWriter(filePath, engine='openpyxl') as writer:
    for i in buy_uniqueLettuce:
        buydf_lettuce_array[j].to_excel(writer, sheet_name=i)
        j += 1


#トマト仕入返品リストの作成
fileName = dayName + 'トマト仕入返品データ.xlsx'
filePath = outputFolderPath+fileName
wb = openpyxl.Workbook()
for i in buy_returneTomato:
    wb.create_sheet(i)
wb.save(filePath)
j = 0
with pd.ExcelWriter(filePath, engine='openpyxl') as writer:
    for i in buy_returneTomato:
        buydf_returneTomato_array[j].to_excel(writer, sheet_name=i)
        j += 1



#レタス仕入返品データベース
fileName = dayName + '仕入レタス売上返品データ.xlsx'
filePath = outputFolderPath+fileName
wb = openpyxl.Workbook()
for i in buy_returneLettuce:
    wb.create_sheet(i)
wb.save(filePath)
j = 0
with pd.ExcelWriter(filePath, engine='openpyxl') as writer:
    for i in buy_returneLettuce:
        buydf_returneLettuce_array[j].to_excel(writer, sheet_name=i)
        j += 1
print('仕入データを作成')
############################仕入Db作成終了####################
