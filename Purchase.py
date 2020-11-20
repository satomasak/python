import pandas as pd
#sekkdf�ɕK�v�ȃJ�����ԍ������̃f�[�^�t���[���𔄏�.xls���璊�o����
selldf = pd.read_excel('C:\\Users\\Himuka\\Desktop\\purchaseData\\����.xls',usecols=[1,5,6,9,11,12,15,17,18])

#procuctRist�̍쐬�@���i�ꗗ�@�d�����폜���A���j�[�N�f�[�^�[�𒊏o
productList = selldf[['����敪','���i��']].drop_duplicates()
#purchaseRist�̍쐬�@�d����ꗗ�@����敪�́u�|���̑��v�Œ��o
purchaseList = productList[productList['����敪'] == '�|���̑�']
#�d����ꗗ�̒�����g�}�g�A���^�X���敪
tomatoPurchaseList = purchaseList[purchaseList['���i��'].str.contains('�g�}�g|����|����|���P|���')]
lettucePurchaseList = purchaseList[purchaseList['���i��'].str.contains('���^�X|���')]

#�ԕi�ȊO�̃g�}�g���i�̃��X�g
x = tomatoPurchaseList [~tomatoPurchaseList ['���i��'].str.contains('�ԕi')]
simpleListT = x['���i��']
#�ԕi�ȊO�̃��^�X�d�����i�̃��X�g
x = lettucePurchaseList [~lettucePurchaseList ['���i��'].str.contains('�ԕi')]
simpleListL = x['���i��']



#�g�}�g�d���ꏤ�i�̔���f�[�^�����i���A���Ӑ�A����P�����ƂɃO���[�v������(�ԕi�ȊO)
tomatoSellData = []
for i in simpleListT:
    nam = selldf[selldf['���i��']== i]
    #���Ӑ�R�[�h�A���Ӑ於�A���i���A����P�����L�[�ɂ��āA��������E���ʂ͍��v����
    nam = nam.groupby(['���Ӑ溰��','���Ӑ於','����N����','���i��','����P��']).sum()
    tomatoSellData.append(nam)

#���^�X�̔���f�[�^�����i���A���Ӑ�A����P�����ƂɃO���[�v��������(�ԕi�ȊO)
lettuceSellData = []
for i in simpleListL:
    nam = selldf[selldf['���i��']== i]
    nam = nam.groupby(['���Ӑ溰��','���Ӑ於','����N����','���i��','����P��']).sum()
    lettuceSellData.append(nam)



#�d����G�N�Z������K�v�ȗ�݂̂�ǂݍ���
buydf = pd.read_excel('C:\\Users\\Himuka\\Desktop\\purchaseData\\�d����.xls',usecols=[1,5,6,9,11,12,15,17,18,21,22])

#�d���ꗗ����g�}�g�̃f�[�^�𒊏o���A�d����E���i�E�d���P�����ƂɃO���[�v������
tomatoBuyData = []
for i in simpleListT:
    nam = buydf[buydf['���i��']== i]
    nam = nam.groupby(['�d���溰��','�d���於','����N����','���i��','�d���P��']).sum()
    tomatoBuyData.append(nam)

#�d����ꗗ����d�����^�X�̃f�[�^�𒊏o���A�d����E���i�E�d���P�����ƂɃO���[�v������
lettuceBuyData = []
for i in simpleListL:
    nam = buydf[buydf['���i��']== i]
    nam = nam.groupby(['�d���溰��','�d���於','����N����','���i��','�d���P��','�s�E�v']).sum()
    lettuceBuyData.append(nam)



import openpyxl
import datetime
filePath = 'C:\\Users\\Himuka\\Desktop\\sellData\\�d�����i����f�[�^.xlsx'
sellData = openpyxl.load_workbook(filePath)



tomatoSheet = sellData['tomato']
i = 3
for nam in simpleListT:
    tomatoSheet.cell(row=i,column=3).value = nam
    i += 1



lettuceSheet = sellData['lettuce']
i = 3
for nam in simpleListL:
    lettuceSheet.cell(row=i,column=3).value = nam
    i += 1


sellData.save(filePath)