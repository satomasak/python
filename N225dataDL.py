#!/usr/bin/env python
# coding: utf-8
#JPXのサイトから特定のファイルをダウンロードし加工するプログラムを作成する
#デリバティブ建玉残高表は毎営業日更新されるため、直近のデータ以外は収集できない。データ加工を自動化できるとDLし忘れがなく効率化が可能
#１．JPXサイトにアクセスし特定のファイルをダウンロードする
#２．ダウンロードディレクトリから任意のディレクトリへファイルを移動する
#3.　デリバティブ建玉残高表から必要なデータを別のエクセルに張り付ける

#自動化　PCの起動、プログラムの実行、シャットダウンはPCのBIOSやタスクスケジューラで設定
#毎日20:30にプログラムを実行するようにタスクスケジューラで設定
#営業日以外はデータ更新が行われないため、if文で営業日であればプログラムを実行するように設定する
#営業日のみ実行するためにjpbizdayのインポート
import jpbizday as biz
#営業日であれば真
if biz.is_bizday(day):

    #１．以下、JPXのサイトに接続し、デリバティブ建玉残高表をダウンロードするためのコード
    from selenium import webdriver
    import time
    #chrome自動操作のためのwebdriberのパスを設定
    #webdriberの取得と設定に関してはキノコードさんのyoutubeを参照
    page01 = webdriver.Chrome(executable_path = 'C:\\Users\\sato\\Desktop\\mypandas\\chromedriver.exe')
    #待機3秒
    page01.implicitly_wait(3)
    # JPXの当日取引高等のページにアクセス
    url01 = "https://www.jpx.co.jp/markets/derivatives/trading-volume/index.html"
    page01.get(url01)
    #待機3秒
    time.sleep(3)
    #OSEデリバティブ建玉残高表のダウンロード　
    #HTMLのボタンにIDやname属性がないためXパスでダウンロードボタンを指定
    #Xパスの取得についてはキノコードさんのyoutubeを参照
    dataDL = page01.find_element_by_xpath('/html/body/div[2]/div[4]/div[1]/div[3]/div[3]/div/table/tbody/tr[2]/td[3]/a')
    time.sleep(1)
    dataDL.click()
    #ダウンロード時間に余裕を持たせるための待機
    time.sleep(10)
    #以上、ソースになるファイルのダウンロード完了



    #２．以下、ダウンロードファイルを任意のディレクトリへ移動するためのコード
    #OSEデリバティブ建玉残高表のファイル名に日付が含まれているため、当日の日付をファイル名に付けるように加工する
    import datetime
    day = datetime.date.today()
    #本日の日付であるdatetime型のデータ(2020-08-28)からハイフンを抜く
    dayName = day.strftime('%Y%m%d')
    #日付＋共通ファイル名
    #OSEデリバティブ建玉残高表は毎営業日の20:00に更新されるため、20:00以前および土日祝日にプログラムを実行するとここでエラーになるので注意
    fileName = dayName+'open_interest.xls'
    #ファイル移動のためのインポート
    import shutil
    #ダウンロードディレクトリから指定ディレクトリへ移動　
    #ここではDownloads⇒datasource　設定に応じて任意のディレクトリへ変更
    shutil.move('C:\\Users\\sato\\Downloads\\'+fileName, 'C:\\Users\\sato\\Desktop\\datasource')
    #以上、ソースになるダウンロードファイルの移動完了



    #３.以下、デリバティブ建玉残高表から必要なデータをvolume&open_contract.xlsx張り付ける
    import openpyxl as px
    #デリバティブ建玉残高表の拡張子が.xlsであるため、excel2003以前のエクセルを操作するためのxlrdライブラリをインポート
    import xlrd
    #デリバティブ建玉残高表の読み込み
    ds = xlrd.open_workbook('C:\\Users\\sato\\Desktop\\datasource\\'+fileName)
    #シートの読み込み
    sheet01 = ds.sheet_by_name('デリバティブ建玉残高状況')
    #貼り付け先エクセルの読み込み
    vc = px.load_workbook('C:\\Users\\sato\\Desktop\\futuresdata\\volume&open_contract.xlsx')
    #シートの読み込み
    sheet02 = vc['vol']
    #貼り付けをする4行目に行を挿入
    sheet02.insert_rows(4)
    #日付のセルに本日の日付を挿入
    sheet02["B4"] = day
    sheet02["L4"] = day
    #取引高のコピー
    sheet02["C4"].value = sheet01.cell(36,2).value  #N225全体
    sheet02["D4"].value = sheet01.cell(17,2).value  #N225期近
    sheet02["E4"].value = sheet01.cell(33,9).value  #N225mini全体
    sheet02["F4"].value = sheet01.cell(17,9).value  #N225mini期近
    sheet02["G4"].value = sheet01.cell(42,2).value  #topix全体
    sheet02["H4"].value = sheet01.cell(37,2).value  #topix期近
    sheet02["I4"].value = sheet01.cell(147,2).value #プット全体
    sheet02["J4"].value = sheet01.cell(148,2).value #コール全体
    #残玉のコピー
    sheet02["M4"].value = sheet01.cell(36,3).value  #N225全体
    sheet02["O4"].value = sheet01.cell(17,3).value  #N225期近
    sheet02["Q4"].value = sheet01.cell(33,10).value #N225mini全体
    sheet02["S4"].value = sheet01.cell(17,10).value #N225mini期近
    sheet02["U4"].value = sheet01.cell(42,3).value  #topix全体
    sheet02["W4"].value = sheet01.cell(37,3).value  #topix期近
    sheet02["Y4"].value = sheet01.cell(147,3).value #プット全体
    sheet02["AA4"].value = sheet01.cell(148,3).value #コール全体
    #残玉の前日比のコピー
    sheet02["N4"].value = sheet01.cell(36,4).value
    sheet02["P4"].value = sheet01.cell(17,4).value
    sheet02["R4"].value = sheet01.cell(33,11).value
    sheet02["T4"].value = sheet01.cell(17,11).value
    sheet02["V4"].value = sheet01.cell(42,4).value
    sheet02["X4"].value = sheet01.cell(37,4).value
    sheet02["Z4"].value = sheet01.cell(147,4).value
    sheet02["AB4"].value = sheet01.cell(148,4).value
    #挿入したセルに罫線を引くためのインポート
    from openpyxl.styles.borders import Border, Side
    #線の指定（細線）
    side = Side(style='thin')
    #罫線を引く場所の指定（右側）
    border = Border(right=side)
    #4行目1から29セルまでの繰り返し処理
    for i in range(1,29):
        sheet02.cell(4, i).border = border
    #貼り付けをしたエクセルの保存
    vc.save('C:\\Users\\sato\\Desktop\\futuresdata\\volume&open_contract.xlsx')
    #以上、データの貼り付け完了
