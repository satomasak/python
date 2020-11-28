import openpyxl as xl

dataPath = 'C:\\Users\\Himuka\\Desktop\\sellData\\'
dbPath = 'C:\\Users\\Himuka\\Desktop\\仕入管理\\'
outPath = 'C:\\Users\\Himuka\\Desktop\\output\\'

otomemini = sellwb['日向おとめミニ']

#表のタイトル作成
wb = openpyxl.Workbook()
sheets = wb.create_sheet('仕入商品利益率')
title = ['商品名','売上単価','売上数','売上金額''仕入単価','仕入数量','仕入金額','粗利','運賃','フィ-','雑費','利益金額','利益率']

j = 2
for i in title:
    sheets.cell(row=2, column=j).value = i
    j += 1
wb.save(outPath+'仕入商品利益率.xlsx')
outFile = outPath+'仕入商品利益率.xlsx'

