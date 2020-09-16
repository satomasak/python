#N225のday、nightの4本値の収集を自動化するプログラム
#＊夜間足、日中足を作成するための情報収取が目的であり、リアルタイムの価格を取得するプログラムではあらいません
#＊リアルタイム価格を取得したい場合は各証券会社が提供しているツールを使用するようにしてください

#使用するライブラリ
#beautifulsoup4、reaquests,openpyxl

#タスクスケジューラはサイトの価格更新時間を踏まえた時間設定をしてください
#N225先物は夜間取引から翌日の日中取引を一日とみなしていますが、日付は日中取引を使用しています
#例えば、9月2日の１日の取引は、9月1日の16:30に開始する夜間取引と9月2日の日中取引の終わり15:30までの間の取引を合わせたものになります
#そのため、夜間取引は翌営業日の日付になるため、注意が必要です
#例えば、金曜日から土曜の早朝までに取引された夜間取引の日付は、翌週月曜日の日付になります。
#夜間取引の価格取得は日中取引が行われる日の16:00くらいまでに時間設定をすると営業日や日付にずれが起きません


#以下、日中取引の四本値を収集するプログラムですが、夜間取引は別シートに記述するように設定しなおせば収集可能です
#営業日のみ実行するためにjpbizdayのインポート
import jpbizday as biz
import datetime
day = datetime.date.today()
#営業日であれば真、データ収集の処理を実行する
if biz.is_bizday(day):
    print("N225のの四本値を取得します")

    #webページ情報を取得するためのrequestsライブラリのインポート
    import requests
    #dayの4本値が表示されているページの取得
    res = requests.get('取得先のURLを記述')
    from bs4 import BeautifulSoup as BS
    #BSのパーサーを使いHTMLの解析を行う
    bsRes = BS(res.text, "html.parser")
    #N225期近の4本値が記載されているHTMLのclass名から値段を取得
    values = bsRes.select("idやclass名を記述")

    #集計用ファイルの読み込み
    import openpyxl as px
    f4val = px.load_workbook('取得したデータをまとめるエクセルのパスを記述')
    #day用のシートの指定
    dayst = f4val['シート名を記述']
    #貼り付けをする4行目に行を挿入
    dayst.insert_rows(4)
    #日付の挿入
    dayVal = bsRes.select("idやclass名を記述")
    dayst['A4'] = dayVal[20].contents[0].strip("\r\t")
    #valuesの0番目の要素（td class="a-right">	23,060<br/>(08:45)	</td>）の0番目のcontents('\r\t\t\t\t\t\t\t\t\t23,060')から
    #\r\tの文字列を削除し、,を空白に置き換えint化する処理。その後集計ファイルの所定箇所に値を代入する処理
    for i in range(0,4):
        dayst.cell(4, i+2).value = int(values[i].contents[0].strip("\r\t").replace(',', ''))
        #for文での処理をわかりやすく記述すると以下の通り
        #dayst['B4'] = int(values[0].contents[0].strip("\r\t").replace(',', ''))    #始値
        #dayst['C4'] = int(values[1].contents[0].strip("\r\t").replace(',', ''))    #高値
        #dayst['D4'] = int(values[2].contents[0].strip("\r\t").replace(',', ''))    #安値
        #dayst['E4'] = int(values[3].contents[0].strip("\r\t").replace(',', ''))    #終値（現在値）
    #挿入したセルに罫線を引くためのインポート
    from openpyxl.styles.borders import Border, Side
    #線の指定（細線）
    side = Side(style='thin')
    #罫線を引く場所の指定（右側, 下）
    border = Border(right=side, bottom=side, left=side)
    #4行目1から5セルまでの繰り返し処理
    for i in range(1,6):
        dayst.cell(4, i).border = border
    #ファイルのセーブ
    f4val.save('取得したデータをまとめるエクセルのパスを記述')

else:print("本日はN225の取引はありません。")
