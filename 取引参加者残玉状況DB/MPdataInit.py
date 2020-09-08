#取引参加者週間残玉状況をエクセルでデータベース化する
#DB.xlsxにの初期データ挿入のためのプログラム
#ソースになるデータはhttps://www.jpx.co.jp/markets/derivatives/open-interest/index.htmlの指数先物取引から取得する
#ここでは、週ごとの取引参加者ごとの残玉変化を追えるようにDBの初期化を行う
#取引参加者一覧（MP一覧）はhttps://www.jpx.co.jp/rules-participants/participants/list/index.htmlから取得


#DB追加の準備
import openpyxl as px
#データソースになる週間取引参加者残玉状況のファイルを読み込む
dsmp = px.load_workbook('C:\\Users\\sato\\Desktop\\datasource\\MP\\20200807_indexfut_oi_by_tp.xlsx')
#シートの指定
dsst = dsmp['BO_DD0120']
#DB用ファイルの読み込み
db = px.load_workbook('C:\\Users\\sato\\Desktop\\futuresdata\\DB.xlsx')
#DBエクセルのシートの指定
N225db = db['N225MP']   #N225の残玉記録シート
N225minidb = db['N225miniMP']   #N225miniの残玉記録シート
topixdb = db['topixMP'] #topixの残玉記録シート
MPdb = db['MP一覧']   #取引参加者一覧シート


#日付の加工
dateds = dsst['A2'].value
y = dateds[1:6]
m = dateds[7:9]
d = dateds[10:12]
date = y+'-'+m+'-'+d

#エクセルのD列に新規データを入れる列を追加する
N225db.insert_cols(4)

#日付と限月を挿入するセルのフォントサイズの設定（セルからあふれるため）
from openpyxl.styles.fonts import Font
N225db['D1'].font = Font(size=10)
N225db['D2'].font = Font(size=8)

#N225MPシートのデータ挿入処理の開始
#日付の代入
N225db['D1'] = date
#限月の代入
N225db['D2'] = dsst['B6'].value

#売り越し参加者のコピー 売り越し玉はマイナス表記で代入する
#上位15位までしかデータはないのでrange()で範囲指定。ただし、参加者が15に満たない場合空白で表記されるため、空白になったらbreakするようにする
for i in range(6,21):
    if dsst.cell(i,3).value == None:
        break
    #会社コードが文字列扱いになっているため数値に変換
    else:
        N225db.cell(i-3, 2).value = int(dsst.cell(i, 3).value)
        #売り越し残玉のため−で表記する
        N225db.cell(i-3, 4).value = -(dsst.cell(i, 5).value)
#売り越しデータの最後から買い越しデータをコピーしていくため、売り越しデータの最後の行をmaxiに代入しておく
maxi = i-3


#買い越し参加者のコピー
for i in range(6,21):
    maxi += 1
    if dsst.cell(i,6).value == None:
        break
    #行は売り越しデータの最後から追加していく
    else:
        N225db.cell(maxi, 2).value = int(dsst.cell(i, 6).value)
        N225db.cell(maxi, 4).value = dsst.cell(i, 8).value


#B列のコードをもとにMP一覧のシートから対応する取引参加者名を検索しC列に代入する処理
#MP一覧から取引参加者は122行まで登録されているので最大範囲は3~123行
for i in range(3, 123):
    #セルの値が空白になったらbreak
    if N225db.cell(i,2).value == None:
        break
    #N225MPのコードとMP一覧のコードが一致したらC列に取引参加者名をコピーする
    else:
        N225code = N225db.cell(i, 2).value
        for im in range(3, 123):
            MPcode = MPdb.cell(im, 2).value
            if N225db.cell(i, 2).value == MPdb.cell(im, 2).value:
                N225db.cell(i, 3).value = MPdb.cell(im, 3).value
#N225の処理終了



#N225miniシートの初期データ挿入処理
N225minidb.insert_cols(4)
#挿入した行のセルにフォント設定
N225minidb['D1'].font = Font(size=10)
N225minidb['D2'].font = Font(size=8)

N225minidb['D1'] = date
N225minidb['D2'] = dsst['B24'].value

#売り越し参加者のコピー
for i in range(24,39):
    if dsst.cell(i,3).value == None:
        break
    else:
        N225minidb.cell(i-21, 2).value = int(dsst.cell(i, 3).value)
        N225minidb.cell(i-21, 4).value = -(dsst.cell(i, 5).value)

maxi = i-21

#買い越し参加者のコピー
for i in range(24,39):
    maxi += 1
    if dsst.cell(i,6).value == None:
        break
    else:
        N225minidb.cell(maxi, 2).value = int(dsst.cell(i, 6).value)
        N225minidb.cell(maxi, 4).value = dsst.cell(i, 8).value


#B列のコードをもとにMP一覧のシートから対応する取引参加者名を検索しC列に代入する処理
#MP一覧から取引参加者は122行まで登録されているので最大範囲は3~123行
for i in range(3, 123):
    #セルの値が空白になったらbreak
    if N225minidb.cell(i,2).value == None:
        break
    #N225miniMPのコードとMP一覧のコードが一致したらC列に取引参加者名をコピーする
    else:
        N225minicode = N225minidb.cell(i, 2).value
        for im in range(3, 123):
            MPcode = MPdb.cell(im, 2).value
            if N225minidb.cell(i, 2).value == MPdb.cell(im, 2).value:
                N225minidb.cell(i, 3).value = MPdb.cell(im, 3).value
#N225miniの処理終了



#topixシートの初期データ挿入処理
topixdb.insert_cols(4)
#挿入した列のセルにフォント設定
topixdb['D1'].font = Font(size=10)
topixdb['D2'].font = Font(size=8)

topixdb['D1'] = date
topixdb['D2'] = dsst['B24'].value


#売り越し参加者のコピー
for i in range(42,57):
    if dsst.cell(i,3).value == None:
        break
    else:
        topixdb.cell(i-39, 2).value = int(dsst.cell(i, 3).value)
        topixdb.cell(i-39, 4).value = -(dsst.cell(i, 5).value)

maxi = i-39


#買い越し参加者のコピー
for i in range(42,57):
    maxi += 1
    if dsst.cell(i,6).value == None:
        break
    else:
        topixdb.cell(maxi, 2).value = int(dsst.cell(i, 6).value)
        topixdb.cell(maxi, 4).value = dsst.cell(i, 8).value


#B列のコードをもとにMP一覧のシートから対応する取引参加者名を検索しC列に代入する処理
#MP一覧から取引参加者は122行まで登録されているので最大範囲は3~123行
for i in range(3, 123):
    #セルの値が空白になったらbreak
    if topixdb.cell(i,2).value == None:
        break
    #topixMPのコードとMP一覧のコードが一致したらC列に取引参加者名をコピーする
    else:
        topixdbcode = topixdb.cell(i, 2).value
        for im in range(3, 123):
            MPcode = MPdb.cell(im, 2).value
            if topixdb.cell(i, 2).value == MPdb.cell(im, 2).value:
                topixdb.cell(i, 3).value = MPdb.cell(im, 3).value

#topixの処理終了

db.save('C:\\Users\\sato\\Desktop\\futuresdata\\DB.xlsx')
#処理後はDB初期化済.xlsxを参照
