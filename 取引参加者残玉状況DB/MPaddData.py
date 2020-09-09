#取引者別残高表にデータを追加するプログラム(毎週月曜日更新)

#データ追加の準備（MPdataInit.pyからコピペ）
import openpyxl as px
#データソースになる週間取引参加者残玉状況のファイルを読み込む
dsmp = px.load_workbook('C:\\Users\\sato\\Desktop\\datasource\\MP\\20200904_indexfut_oi_by_tp.xlsx')
#シートの指定
dsst = dsmp['BO_DD0120']
#DB用ファイルの読み込み
db = px.load_workbook('C:\\Users\\sato\\Desktop\\futuresdata\\DB.xlsx')
#DBエクセルのシートの指定
N225db = db['N225MP']   #N225の残玉記録シート
N225minidb = db['N225miniMP']   #N225miniの残玉記録シート
topixdb = db['topixMP'] #topixの残玉記録シート
MPdb = db['MP一覧']   #取引参加者一覧シート

#エクセルのD列に新規データを入れる列を追加する
N225db.insert_cols(4)
N225minidb.insert_cols(4)
topixdb.insert_cols(4)

#日付と限月を挿入するセルのフォントサイズの設定（セルからあふれるため）
from openpyxl.styles.fonts import Font
N225db['D1'].font = Font(size=10)
N225db['D2'].font = Font(size=8)
N225minidb['D1'].font = Font(size=10)
N225minidb['D2'].font = Font(size=8)
topixdb['D1'].font = Font(size=10)
topixdb['D2'].font = Font(size=8)

#日付の加工
dateds = dsst['A2'].value
y = dateds[1:6]
m = dateds[7:9]
d = dateds[10:12]
date = y+'-'+m+'-'+d

#日付と限月の代入
N225db['D1'] = date
N225db['D2'] = dsst['B6'].value
N225minidb['D1'] = date
N225minidb['D2'] = dsst['B24'].value
topixdb['D1'] = date
topixdb['D2'] = dsst['B24'].value
#データ追加の準備完了


#N225のデータ追加の処理
#シートのデータがある最大行数を調べる
#max_rowを使うと書式設定が変更されていると空白セルでもカウントされてしまうためエラーを起こさないためにカウンターで調べる
#行数の先頭の数字を変数rowCountに代入
rowCount = 3
#rowCountにデータが入っている最後の行＋１のデータが入っている
while N225db.cell(rowCount, 2).value != None:
    rowCount += 1


#N225の残玉を代入する処理

#N225　売り越しデータ処理
#dsi:データソース用変数　dbi：DB用変数　mpi：MP一覧用変数
for dsi in range(6, 21):
    #データソースファイルに空白があれば処理を抜ける
    if dsst.cell(dsi,3).value == None:
        break
    else:
        #データソースのファイルからコード番号をint型で変数に代入
        N225code = int(dsst.cell(dsi, 3).value)
        #N225MP B列のコードと一致する残玉データを、D列に代入する処理
        #N225codeと一致せずに最後の行まで検索をしたならば、新規登録をする
        for dbi in range(3, rowCount):
            #すべての行を検索するため、検索が最後の行まで行われた場合、取引者の新規登録を行うようにする
            if N225db.cell(dbi, 2).value != N225code:
                if dbi == rowCount-1:

                    #新規登録処理
                    print('2020-8-7以降、N225を取引していなかった参加者が見つかったので情報を追加します')
                    #コードと残玉を最後の行に追加する
                    N225db.cell(rowCount, 2).value = N225code
                    N225db.cell(rowCount, 4).value = -(dsst.cell(dsi, 5).value)

                    #コードをもとにMP一覧から会社名を検索し代入する 検索が最後の行まで行われた場合、MP一覧に登録されていない取引者が見つかったことをprintする
                    for mpi in range(3, 123):
                        if MPdb.cell(mpi, 2).value != N225code:
                            if mpi == 122:
                                print('MP一覧に登録されていないと取引者が見つかりました コード番号：'+str(N225code))
                                break
                        else:
                            N225db.cell(rowCount, 3).value = MPdb.cell(mpi, 3).value
                            rowCount += 1
                            break

            #コードが一致するものは残玉データを代入していく
            else:
                N225db.cell(dbi, 4).value = -(dsst.cell(dsi, 5).value)
                break


#N225　買い越しデータ処理
for dsi in range(6, 21):
    #データソースファイルに空白があれば処理を抜ける
    if dsst.cell(dsi,6).value == None:
        break
    else:
        #データソースのファイルからコード番号をint型で変数に代入
        N225code = int(dsst.cell(dsi, 6).value)
        #N225MP B列のコードと一致する残玉データを、D列に代入する処理
        #N225codeと一致せずに最後の行まで検索をしたならば、新規登録をする
        for dbi in range(3, rowCount):
            #すべての行を検索するため、検索が最後の行まで行われた場合、取引者の新規登録を行うようにする
            if N225db.cell(dbi, 2).value != N225code:
                if dbi == rowCount-1:

                    #新規登録処理
                    print('2020-8-7以降、N225を取引していなかった参加者が見つかったので情報を追加します')
                    #コードと残玉を最後の行に追加する
                    N225db.cell(rowCount, 2).value = N225code
                    N225db.cell(rowCount, 4).value = dsst.cell(dsi, 8).value

                    #コードをもとにMP一覧から会社名を検索し代入する 検索が最後の行まで行われた場合、MP一覧に登録されていない取引者が見つかったことをprintする
                    for mpi in range(3, 123):
                        if MPdb.cell(mpi, 2).value != N225code:
                            if mpi == 122:
                                print('MP一覧に登録されていないと取引者が見つかりました コード番号：'+str(N225code))
                                break
                        else:
                            N225db.cell(rowCount, 3).value = MPdb.cell(mpi, 3).value
                            rowCount += 1
                            break

            #コードが一致するものは残玉データを代入していく
            else:
                N225db.cell(dbi, 4).value = dsst.cell(dsi, 8).value
                break

#N225miniシートのデータがある最大行数を調べる
#max_rowを使うと書式設定が変更されていると空白セルでもカウントされてしまうためエラーを起こさないためにカウンターで調べる
#行数の先頭の数字を変数rowCountに代入
rowCount = 3
#rowCountにデータが入っている最後の行＋１のデータが入っている
while N225minidb.cell(rowCount, 2).value != None:
    rowCount += 1

#N225miniの残玉を代入する処理

#N225mini　売り越しデータ処理
for dsi in range(24, 39):
    #データソースファイルに空白があれば処理を抜ける
    if dsst.cell(dsi,3).value == None:
        break
    else:
        #データソースのファイルからコード番号をint型で変数に代入
        N225minicode = int(dsst.cell(dsi, 3).value)
        #N225miniMP B列のコードと一致する残玉データを、D列に代入する処理
        #N225minicodeと一致せずに最後の行まで検索をしたならば、新規登録をする
        for dbi in range(3, rowCount):
            #すべての行を検索するため、検索が最後の行まで行われた場合、取引者の新規登録を行うようにする
            if N225minidb.cell(dbi, 2).value != N225minicode:
                if dbi == rowCount-1:

                    #新規登録処理
                    print('2020-8-7以降、N225miniを取引していなかった参加者が見つかったので情報を追加します')
                    #コードと残玉を最後の行に追加する
                    N225minidb.cell(rowCount, 2).value = N225minicode
                    N225minidb.cell(rowCount, 4).value = -(dsst.cell(dsi, 5).value)

                    #コードをもとにMP一覧から会社名を検索し代入する 検索が最後の行まで行われた場合、MP一覧に登録されていない取引者が見つかったことをprintする
                    for mpi in range(3, 123):
                        if MPdb.cell(mpi, 2).value != N225minicode:
                            if mpi == 122:
                                print('MP一覧に登録されていないと取引者が見つかりました コード番号：'+str(N225minicode))
                                break
                        else:
                            N225minidb.cell(rowCount, 3).value = MPdb.cell(mpi, 3).value
                            rowCount += 1
                            break

            #コードが一致するものは残玉データを代入していく
            else:
                N225minidb.cell(dbi, 4).value = -(dsst.cell(dsi, 5).value)
                break

#N225mini　買い越しデータ処理
for dsi in range(24, 39):
    #データソースファイルに空白があれば処理を抜ける
    if dsst.cell(dsi,6).value == None:
        break
    else:
        #データソースのファイルからコード番号をint型で変数に代入
        N225minicode = int(dsst.cell(dsi, 6).value)
        #N225miniMP B列のコードと一致する残玉データを、D列に代入する処理
        #N225minicodeと一致せずに最後の行まで検索をしたならば、新規登録をする
        for dbi in range(3, rowCount):
            #すべての行を検索するため、検索が最後の行まで行われた場合、取引者の新規登録を行うようにする
            if N225minidb.cell(dbi, 2).value != N225minicode:
                if dbi == rowCount-1:

                    #新規登録処理
                    print('2020-8-7以降、N225miniを取引していなかった参加者が見つかったので情報を追加します')
                    #コードと残玉を最後の行に追加する
                    N225minidb.cell(rowCount, 2).value = N225minicode
                    N225minidb.cell(rowCount, 4).value = dsst.cell(dsi, 8).value

                    #コードをもとにMP一覧から会社名を検索し代入する 検索が最後の行まで行われた場合、MP一覧に登録されていない取引者が見つかったことをprintする
                    for mpi in range(3, 123):
                        if MPdb.cell(mpi, 2).value != N225minicode:
                            if mpi == 122:
                                print('MP一覧に登録されていないと取引者が見つかりました コード番号：'+str(N225minicode))
                                break
                        else:
                            N225minidb.cell(rowCount, 3).value = MPdb.cell(mpi, 3).value
                            rowCount += 1
                            break

            #コードが一致するものは残玉データを代入していく
            else:
                N225minidb.cell(dbi, 4).value = dsst.cell(dsi, 8).value
                break


#topixの残玉処理

#topixシートのデータがある最大行数を調べる
#行数の先頭の数字を変数rowCountに代入
rowCount = 3
#rowCountにデータが入っている最後の行＋１のデータが入っている
while topixdb.cell(rowCount, 2).value != None:
    rowCount += 1

#topixdbの残玉を代入する処理

#topix　売り越しデータ処理
for dsi in range(42, 57):
    #データソースファイルに空白があれば処理を抜ける
    if dsst.cell(dsi,3).value == None:
        break
    else:
        #データソースのファイルからコード番号をint型で変数に代入
        topixcode = int(dsst.cell(dsi, 3).value)
        #topixdbMP B列のコードと一致する残玉データを、D列に代入する処理
        #topixcodeと一致せずに最後の行まで検索をしたならば、新規登録をする
        for dbi in range(3, rowCount):
            #すべての行を検索するため、検索が最後の行まで行われた場合、取引者の新規登録を行うようにする
            if topixdb.cell(dbi, 2).value != topixcode:
                if dbi == rowCount-1:

                    #新規登録処理
                    print('2020-8-7以降、topixを取引していなかった参加者が見つかったので情報を追加します')
                    #コードと残玉を最後の行に追加する
                    topixdb.cell(rowCount, 2).value = topixcode
                    topixdb.cell(rowCount, 4).value = -(dsst.cell(dsi, 5).value)

                    #コードをもとにMP一覧から会社名を検索し代入する 検索が最後の行まで行われた場合、MP一覧に登録されていない取引者が見つかったことをprintする
                    for mpi in range(3, 123):
                        if MPdb.cell(mpi, 2).value != topixcode:
                            if mpi == 122:
                                print('MP一覧に登録されていないと取引者が見つかりました コード番号：'+str(topixcode))
                                break
                        else:
                            topixdb.cell(rowCount, 3).value = MPdb.cell(mpi, 3).value
                            rowCount += 1
                            break

            #コードが一致するものは残玉データを代入していく
            else:
                topixdb.cell(dbi, 4).value = -(dsst.cell(dsi, 5).value)
                break

#topix　買い越しデータ処理
for dsi in range(42, 57):
    #データソースファイルに空白があれば処理を抜ける
    if dsst.cell(dsi,6).value == None:
        break
    else:
        #データソースのファイルからコード番号をint型で変数に代入
        topixcode = int(dsst.cell(dsi, 6).value)
        #topixMP B列のコードと一致する残玉データを、D列に代入する処理
        #topixcodeと一致せずに最後の行まで検索をしたならば、新規登録をする
        for dbi in range(3, rowCount):
            #すべての行を検索するため、検索が最後の行まで行われた場合、取引者の新規登録を行うようにする
            if topixdb.cell(dbi, 2).value != topixcode:
                if dbi == rowCount-1:

                    #新規登録処理
                    print('2020-8-7以降、topixを取引していなかった参加者が見つかったので情報を追加します')
                    #コードと残玉を最後の行に追加する
                    topixdb.cell(rowCount, 2).value = topixcode
                    topixdb.cell(rowCount, 4).value = dsst.cell(dsi, 8).value

                    #コードをもとにMP一覧から会社名を検索し代入する 検索が最後の行まで行われた場合、MP一覧に登録されていない取引者が見つかったことをprintする
                    for mpi in range(3, 123):
                        if MPdb.cell(mpi, 2).value != topixcode:
                            if mpi == 122:
                                print('MP一覧に登録されていないと取引者が見つかりました コード番号：'+str(topixcode))
                                break
                        else:
                            topixdb.cell(rowCount, 3).value = MPdb.cell(mpi, 3).value
                            rowCount += 1
                            break

            #コードが一致するものは残玉データを代入していく
            else:
                topixdb.cell(dbi, 4).value = dsst.cell(dsi, 8).value
                break

db.save('C:\\Users\\sato\\Desktop\\futuresdata\\DB.xlsx')
